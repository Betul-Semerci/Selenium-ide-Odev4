from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait 
from selenium.webdriver.support import expected_conditions as ec 
import pytest
import openpyxl
from constants import globalConstants as c

class Test_Sauce_Odev2:
    def setup_method(self): 
        self.driver = webdriver.Chrome()
        self.driver.get(c.BASE_URL)
        self.driver.maximize_window() 

    def teardown_method(self): 
        self.driver.quit()
    
    def getData_invalid_login1():
        excel = openpyxl.load_workbook(c.invalid_login_xlsx)
        sheet = excel["Sayfa1"] #hangi sayfada çalışacağımı gösteriyorum
        rows = sheet.max_row #kaçıncı satıra kadar veri var?
        data = []
        for i in range(2,3): #2.satır
            username = sheet.cell(i,1).value #1.sütun
            password = sheet.cell(i,2).value #2.sütun
            data.append((username,password))
        return data
    
    def getData_invalid_login2():
        excel = openpyxl.load_workbook(c.invalid_login_xlsx)
        sheet = excel["Sayfa1"] 
        rows = sheet.max_row 
        data = []
        for i in range(3,4): 
            username = sheet.cell(i,1).value 
            password = sheet.cell(i,2).value 
            data.append((username,password))
        return data
       
    def getData_invalid_login3():
        excel = openpyxl.load_workbook(c.invalid_login_xlsx)
        sheet = excel["Sayfa1"] 
        rows = sheet.max_row 
        data = []
        for i in range(4,rows): 
            username = sheet.cell(i,1).value 
            password = sheet.cell(i,2).value 
            data.append((username,password))
        return data

    def getData_valid_login():
        excel = openpyxl.load_workbook(c.invalid_login_xlsx)
        sheet = excel["Sayfa1"] 
        rows = sheet.max_row 
        data = []
        for i in range(5,rows+1): 
            username = sheet.cell(i,1).value 
            password = sheet.cell(i,2).value 
            data.append((username,password))
        return data  
    

    
    def test_invalid_login1(self):
        """ usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys(username)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys(password) """
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        errorMessage = self.driver.find_element(By.XPATH,c.ERROR_MESSAGE_XPATH1)
        assert errorMessage.text == c.USERNAME_PASSWORD_DONT_MATCH1

    
    
    @pytest.mark.parametrize("username,password",getData_invalid_login2())
    def test_invalid_login2(self,username,password):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys(username)
        """ passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys(password) """
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        errorMessage = self.driver.find_element(By.XPATH,c.ERROR_MESSAGE_XPATH2)
        assert errorMessage.text == c.USERNAME_PASSWORD_DONT_MATCH2

    
    @pytest.mark.parametrize("username,password",getData_invalid_login3())
    def test_invalid_login3(self,username,password):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys(username)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys(password)
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        errorMessage = self.driver.find_element(By.XPATH,c.ERROR_MESSAGE_XPATH3)
        assert errorMessage.text == c.USERNAME_PASSWORD_DONT_MATCH3
    
    
    @pytest.mark.parametrize("username,password",getData_valid_login())
    def test_valid_login(self,username,password):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys(username)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys(password)
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        inventory_items = self.driver.find_elements(By.CLASS_NAME,c.CLASS_NAMEE)
        assert len(inventory_items) == c.LEN
